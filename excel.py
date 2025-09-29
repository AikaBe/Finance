import pandas as pd
from sqlalchemy import create_engine
import os
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

engine = create_engine("postgresql+psycopg2://postgres:080116bs@localhost:5432/finance_db")

df = pd.read_sql("""
    SELECT 
        TO_CHAR(t.date, 'YYYY-MM') AS month,
        SUM(t.amount) AS total_amount
    FROM trans t
    GROUP BY TO_CHAR(t.date, 'YYYY-MM')
    ORDER BY month
""", engine)

df['month'] = df['month'].astype(str)

df['min_total'] = df['total_amount'].min()
df['max_total'] = df['total_amount'].max()


def export_to_excel(dataframes_dict, filename):
    export_path = os.path.join(os.getcwd(), "exports")
    os.makedirs(export_path, exist_ok=True)
    
    full_path = os.path.join(export_path, filename)
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    wb = load_workbook(full_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        ws.freeze_panes = "A2"
        
        ws.auto_filter.ref = ws.dimensions
        
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if all(isinstance(cell.value, (int, float)) or cell.value is None for cell in col):
                col_letter = get_column_letter(col[0].column)
                
                gradient_rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", gradient_rule)
    
    wb.save(full_path)
    
    total_rows = sum(len(df) for df in dataframes_dict.values())
    total_sheets = len(dataframes_dict)
    print(f"Создан файл {filename}, {total_sheets} лист(а/ов), {total_rows} строк")

dataframes = {"Transactions": df}
export_to_excel(dataframes, "transactions_report.xlsx")
